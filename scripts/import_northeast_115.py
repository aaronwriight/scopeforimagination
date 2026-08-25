#!/usr/bin/env python3
"""Import Aaron's Northeast 115 workbook into Venture's peak catalog.

The workbook is the source of truth for public peak facts, completion order,
ratings, ascent counts, dates, and trips. Narrative ascent columns stay private
unless ``--include-private-notes`` is explicitly supplied.

Coordinates can be refreshed from a saved Wilderlist Northeast 111 response.
When that response is omitted, the importer reuses coordinates, stable slugs,
Peakbagger URLs, and linked entry slugs from the existing catalog.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PEAKBAGGER_LIST_URL = "https://www.peakbagger.com/list.aspx?lid=511"
COORDINATE_SOURCE_URL = "https://www.wilderlist.app/list/5db9de782d4ef1001786a442"
DEFAULT_CATALOG_PATH = Path("content/venture/trails/northeast-115.json")
STATE_ABBREVIATIONS = {
    "Maine": "ME",
    "New Hampshire": "NH",
    "New York": "NY",
    "Vermont": "VT",
}
COORDINATE_ALIASES = {
    "Katahdin": "Mount Katahdin - Baxter Peak",
    "Hamlin Peak": "Mount Katahdin - Hamlin Peak",
    "Wildcat Mountain": "Wildcat Mountain, A Peak",
    "Mount Mansfield": "Mount Mansfield - The Chin",
    "Old Speck": "Old Speck Mountain",
    "Mount Osceola - East Peak": "East Osceola",
    "Avery Peak": "Bigelow Mountain - Avery Peak",
    "Wildcat D": "Wildcat Mountain, D Peak",
}
ASCENT_SLOTS = ((1, "first"), (2, "second"), (3, "third"))
REQUIRED_HEADERS = {
    "peak",
    "rank",
    "elev_ft",
    "state",
    "range_level_5",
    "prom_ft",
    "ascents",
    "bagged",
    "ascents_personal",
    "no_completed",
    "rating",
    *(f"date_{label}_ascent" for _, label in ASCENT_SLOTS),
    *(f"trip_{label}_ascent" for _, label in ASCENT_SLOTS),
}


def normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    value = value.replace("&", " and ").replace("'", "")
    value = re.sub(r"\([^)]*\)", " ", value)
    value = value.replace("formerly east dix", "east dix")
    value = value.replace("bondcliffs", "bondcliff")
    value = value.replace("table top", "tabletop")
    value = value.replace("rocky peak ridge", "rocky peak")
    value = value.replace("camels hump", "camel s hump")
    value = re.sub(r"\bmt\.?\b", "mount", value)
    value = re.sub(r"\bmountain\b", "mount", value)
    value = re.sub(r"\bpeak\b", "", value)
    value = re.sub(r"\bthe\b", "", value)
    return " ".join(re.findall(r"[a-z0-9]+", value))


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    value = re.sub(r"\s*\(formerly east dix\)\s*", "", value)
    return re.sub(r"^-|-$", "", re.sub(r"[^a-z0-9]+", "-", value))


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).replace("\u00a0", " ").strip().lower()
    return re.sub(r"^_|_$", "", re.sub(r"[^a-z0-9]+", "_", text))


def find_headers(sheet: Any) -> tuple[int, dict[str, int]]:
    """Find the header row and return canonical header names mapped to columns."""
    for row_number in range(1, min(sheet.max_row, 25) + 1):
        headers = {
            normalize_header(sheet.cell(row_number, column).value): column
            for column in range(1, sheet.max_column + 1)
            if normalize_header(sheet.cell(row_number, column).value)
        }
        if REQUIRED_HEADERS.issubset(headers):
            return row_number, headers

    raise ValueError(
        "Could not find the updated Northeast 115 header row; missing required "
        f"headers from {sorted(REQUIRED_HEADERS)}"
    )


def load_coordinates(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text())
    try:
        peak_list = payload["data"]["peakList"]
        coordinates = peak_list["mountains"] + peak_list.get("optionalMountains", [])
    except (KeyError, TypeError) as error:
        raise ValueError("Coordinate JSON is not a Wilderlist peak-list response") from error

    if len(coordinates) != 115:
        raise ValueError(f"Expected 115 coordinate records, found {len(coordinates)}")
    return coordinates


def coordinate_for(
    peak_name: str,
    state: str,
    coordinates: list[dict[str, Any]],
) -> tuple[float, float]:
    target_name = COORDINATE_ALIASES.get(peak_name, peak_name)
    state_abbreviation = STATE_ABBREVIATIONS[state]
    matches = [
        item
        for item in coordinates
        if item["locationTextShort"] == state_abbreviation
        and normalize_name(item["name"]) == normalize_name(target_name)
    ]
    if len(matches) != 1:
        raise ValueError(
            f"Expected one coordinate match for {peak_name}, found "
            f"{len(matches)}: {[item['name'] for item in matches]}"
        )

    longitude, latitude = matches[0]["location"]
    return round(float(latitude), 7), round(float(longitude), 7)


def nonempty_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def integer_value(value: Any, *, field: str, row: int, required: bool = False) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError(f"Missing {field} on workbook row {row}")
        return None
    if isinstance(value, bool):
        raise ValueError(f"Invalid {field} on workbook row {row}: {value!r}")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Invalid {field} on workbook row {row}: {value!r}") from error
    if not number.is_integer():
        raise ValueError(f"Invalid {field} on workbook row {row}: {value!r}")
    return int(number)


def iso_date(value: Any, *, row: int, field: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text or text.casefold() in {"na", "n/a"}:
        return None
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"Invalid {field} on workbook row {row}: {value!r}")


def load_existing_catalog(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    payload = json.loads(path.read_text())
    if not isinstance(payload.get("peaks"), list) or len(payload["peaks"]) != 115:
        raise ValueError(f"Existing catalog at {path} does not contain 115 peaks")
    return payload


def peak_identity(name: str, state: str) -> tuple[str, str]:
    return normalize_name(name), state


def existing_entry_slug(existing_peak: dict[str, Any] | None, ordinal: int) -> str | None:
    if not existing_peak:
        return None
    for ascent in existing_peak.get("ascents", []):
        if ascent.get("ordinal") == ordinal:
            value = ascent.get("entrySlug")
            return value if isinstance(value, str) and value else None
    return None


def build_catalog(
    workbook_path: Path,
    coordinate_path: Path | None = None,
    *,
    existing_catalog_path: Path | None = DEFAULT_CATALOG_PATH,
    include_private_notes: bool = False,
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    header_row, headers = find_headers(sheet)
    coordinates = load_coordinates(coordinate_path) if coordinate_path else None
    existing_catalog = load_existing_catalog(existing_catalog_path)
    if coordinates is None and existing_catalog is None:
        raise ValueError(
            "A Wilderlist response or an existing 115-peak catalog is required for coordinates"
        )

    existing_by_identity = {
        peak_identity(peak["name"], peak["state"]): peak
        for peak in (existing_catalog or {}).get("peaks", [])
    }

    source_rows: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        rank = sheet.cell(row_number, headers["rank"]).value
        name = nonempty_text(sheet.cell(row_number, headers["peak"]).value)
        if not isinstance(rank, (int, float)) or not name:
            continue

        state = nonempty_text(sheet.cell(row_number, headers["state"]).value)
        if state not in STATE_ABBREVIATIONS:
            raise ValueError(f"Invalid state on workbook row {row_number}: {state!r}")

        peak_cell = sheet.cell(row_number, headers["peak"])
        source_rows.append(
            {
                "row": row_number,
                "rank": integer_value(rank, field="rank", row=row_number, required=True),
                "name": name,
                "elevationFeet": integer_value(
                    sheet.cell(row_number, headers["elev_ft"]).value,
                    field="elevation",
                    row=row_number,
                    required=True,
                ),
                "state": state,
                "range": nonempty_text(sheet.cell(row_number, headers["range_level_5"]).value),
                "prominenceFeet": integer_value(
                    sheet.cell(row_number, headers["prom_ft"]).value,
                    field="prominence",
                    row=row_number,
                    required=True,
                ),
                "peakbaggerAscents": integer_value(
                    sheet.cell(row_number, headers["ascents"]).value,
                    field="Peakbagger ascents",
                    row=row_number,
                    required=True,
                ),
                "bagged": nonempty_text(sheet.cell(row_number, headers["bagged"]).value) is not None,
                "timesHiked": integer_value(
                    sheet.cell(row_number, headers["ascents_personal"]).value,
                    field="personal ascents",
                    row=row_number,
                ),
                "completionNumber": integer_value(
                    sheet.cell(row_number, headers["no_completed"]).value,
                    field="completion number",
                    row=row_number,
                ),
                "rating": sheet.cell(row_number, headers["rating"]).value,
                "ascentSlots": [
                    {
                        "ordinal": ordinal,
                        "date": iso_date(
                            sheet.cell(row_number, headers[f"date_{label}_ascent"]).value,
                            row=row_number,
                            field=f"{label} ascent date",
                        ),
                        "trip": nonempty_text(
                            sheet.cell(row_number, headers[f"trip_{label}_ascent"]).value
                        ),
                        "note": (
                            nonempty_text(sheet.cell(row_number, headers[f"{label}_ascent"]).value)
                            if include_private_notes and f"{label}_ascent" in headers
                            else None
                        ),
                    }
                    for ordinal, label in ASCENT_SLOTS
                ],
                "workbookSourceUrl": (
                    peak_cell.hyperlink.target if peak_cell.hyperlink is not None else None
                ),
            }
        )

    if len(source_rows) != 115:
        raise ValueError(f"Expected 115 workbook rows, found {len(source_rows)}")

    base_slugs = [slugify(row["name"]) for row in source_rows]
    duplicate_slugs = {slug for slug, count in Counter(base_slugs).items() if count > 1}
    peaks: list[dict[str, Any]] = []

    for row, base_slug in zip(source_rows, base_slugs, strict=True):
        existing_peak = existing_by_identity.get(peak_identity(row["name"], row["state"]))
        if existing_peak:
            slug = existing_peak["slug"]
        else:
            slug = base_slug
            if slug in duplicate_slugs:
                slug = f"{slug}-{STATE_ABBREVIATIONS[row['state']].lower()}"

        workbook_count = row["timesHiked"]
        workbook_completed = bool(
            row["bagged"] or (workbook_count is not None and workbook_count > 0) or row["completionNumber"]
        )
        # A completed peak may have an unknown ascent count. Keep that absence
        # explicit instead of inventing an ascent or field-note record.
        times_hiked = (
            workbook_count if workbook_count is not None else (None if workbook_completed else 0)
        )
        if times_hiked is not None and times_hiked < 0:
            raise ValueError(f"Negative ascent count on workbook row {row['row']}")

        populated_slots = [
            slot["ordinal"]
            for slot in row["ascentSlots"]
            if slot["date"] is not None or slot["trip"] is not None or slot["note"] is not None
        ]
        if populated_slots and (times_hiked is None or max(populated_slots) > times_hiked):
            raise ValueError(
                f"Workbook row {row['row']} supplies ascent {max(populated_slots)} "
                f"but ascents_personal is {times_hiked}"
            )

        slots_by_ordinal = {slot["ordinal"]: slot for slot in row["ascentSlots"]}
        ascents = []
        for ordinal in range(1, (times_hiked or 0) + 1):
            slot = slots_by_ordinal.get(ordinal, {})
            ascents.append(
                {
                    "ordinal": ordinal,
                    "date": slot.get("date"),
                    "trip": slot.get("trip"),
                    "note": slot.get("note") if include_private_notes else None,
                    "entrySlug": existing_entry_slug(existing_peak, ordinal),
                }
            )

        if coordinates is not None:
            latitude, longitude = coordinate_for(row["name"], row["state"], coordinates)
        elif existing_peak is not None:
            latitude = float(existing_peak["latitude"])
            longitude = float(existing_peak["longitude"])
        else:
            raise ValueError(f"No saved coordinates for {row['name']}, {row['state']}")

        if coordinates is None and existing_peak is not None:
            source_url = existing_peak["sourceUrl"]
        else:
            source_url = row["workbookSourceUrl"] or (
                existing_peak.get("sourceUrl") if existing_peak else None
            )
        if not source_url:
            raise ValueError(f"No Peakbagger source URL for {row['name']}")

        rating = row["rating"]
        peaks.append(
            {
                "rank": row["rank"],
                "slug": slug,
                "name": row["name"],
                "elevationFeet": row["elevationFeet"],
                "state": row["state"],
                "stateAbbreviation": STATE_ABBREVIATIONS[row["state"]],
                "range": row["range"],
                "prominenceFeet": row["prominenceFeet"],
                "peakbaggerAscents": row["peakbaggerAscents"],
                "completed": workbook_completed,
                "completionNumber": row["completionNumber"],
                "rating": float(rating) if rating is not None else None,
                "timesHiked": times_hiked,
                "ascents": ascents,
                "latitude": latitude,
                "longitude": longitude,
                "sourceUrl": source_url,
            }
        )

    completion_numbers = [
        peak["completionNumber"] for peak in peaks if peak["completionNumber"] is not None
    ]
    if len(completion_numbers) != len(set(completion_numbers)):
        raise ValueError("Workbook completion numbers must be unique")
    for peak in peaks:
        if peak["completed"] != (peak["completionNumber"] is not None):
            raise ValueError(
                f"Completion status and completion number disagree for {peak['name']}"
            )

    return {
        "$schema": "./northeast-115.schema.json",
        "name": "Northeast 115",
        "description": "A summit-by-summit log of the 4000-footers across New Hampshire, New York, Maine, and Vermont.",
        "peakbaggerListUrl": PEAKBAGGER_LIST_URL,
        "coordinateSourceUrl": COORDINATE_SOURCE_URL,
        "peaks": peaks,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "coordinates",
        type=Path,
        nargs="?",
        help="Optional saved Wilderlist response; omit to preserve existing coordinates.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_CATALOG_PATH,
    )
    parser.add_argument(
        "--base-catalog",
        type=Path,
        help="Catalog whose stable slugs, coordinates, source URLs, and entry links should be preserved.",
    )
    parser.add_argument(
        "--include-private-notes",
        action="store_true",
        help="Explicitly copy private workbook notes into the public catalog.",
    )
    args = parser.parse_args()

    base_catalog = args.base_catalog
    if base_catalog is None:
        base_catalog = args.output if args.output.exists() else DEFAULT_CATALOG_PATH

    catalog = build_catalog(
        args.workbook,
        args.coordinates,
        existing_catalog_path=base_catalog,
        include_private_notes=args.include_private_notes,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(catalog, indent=2, ensure_ascii=False) + "\n")

    complete = sum(peak["completed"] for peak in catalog["peaks"])
    print(f"Wrote {len(catalog['peaks'])} peaks ({complete} complete) to {args.output}")


if __name__ == "__main__":
    main()
